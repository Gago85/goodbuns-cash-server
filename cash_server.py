from flask import Flask, request, jsonify
from flask_cors import CORS
from datetime import datetime
import os
import json
import requests
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from pathlib import Path
from dotenv import load_dotenv

# 🔐 Загрузка конфигурации
load_dotenv()
TELEGRAM_TOKEN = os.getenv("TELEGRAM_TOKEN")
CHAT_ID = os.getenv("CHAT_ID")
BASE_FOLDER = Path("D:/GoodbunsAIcash")

app = Flask(__name__)
CORS(app)
logging.basicConfig(filename="errors.log", level=logging.ERROR, format="%(asctime)s - %(message)s")

# 💾 Создание красивого отчёта Excel
def create_excel(data, save_folder):
    wb = Workbook()
    ws = wb.active
    ws.title = "Кассовый отчёт"

    # Заголовок
    ws.merge_cells("A1:B1")
    ws["A1"] = "GOODBUNS — КАССОВЫЙ ОТЧЁТ"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Инфо
    ws["A2"] = "Точка:"
    ws["B2"] = data.get("point", "")
    ws["A3"] = "Дата:"
    ws["B3"] = data.get("date", "")

    # Таблица
    ws["A5"] = "Показатель"
    ws["B5"] = "Сумма (₽)"
    ws["A5"].font = ws["B5"].font = Font(bold=True)
    ws["A5"].alignment = ws["B5"].alignment = Alignment(horizontal="center")

    rows = [
        ("Наличные", data.get("cash", 0)),
        ("Безналичные", data.get("card", 0)),
        ("Возврат (нал)", data.get("return_cash", 0)),
        ("Возврат (безнал)", data.get("return_card", 0)),
        ("🧾 Итого", data.get("total", 0)),
        ("🍽 Обеды", data.get("lunches", 0)),
        ("♻️ Списание", data.get("writeoff", 0)),
    ]

    for i, (label, value) in enumerate(rows, start=6):
        ws[f"A{i}"] = label
        ws[f"B{i}"] = float(value)
        ws[f"A{i}"].alignment = Alignment(horizontal="left")
        ws[f"B{i}"].alignment = Alignment(horizontal="right")

    filename = f"Кассовый отчёт - {data['point']} - {data['date']}.xlsx"
    filepath = save_folder / filename
    wb.save(filepath)
    return filepath, filename

def save_json(data, save_folder, filename):
    json_path = save_folder / filename.replace(".xlsx", ".json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def send_to_telegram(filepath):
    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendDocument"
    try:
        with open(filepath, "rb") as file:
            response = requests.post(url, data={
                "chat_id": CHAT_ID,
                "caption": "💰 Новый кассовый отчёт"
            }, files={"document": file})
        logging.error(f"Telegram response: {response.status_code} - {response.text}")
        return response.status_code == 200
    except Exception as e:
        logging.error(f"Telegram send error: {str(e)}")
        return False

@app.route("/submit_cash", methods=["POST"])
def handle_cash():
    try:
        data = request.form.to_dict()
        if not data.get("date"):
            data["date"] = datetime.now().strftime("%Y-%m-%d")

        date_folder = BASE_FOLDER / data["date"]
        date_folder.mkdir(parents=True, exist_ok=True)

        filepath, filename = create_excel(data, date_folder)
        save_json(data, date_folder, filename)
        sent = send_to_telegram(filepath)

        return jsonify({
            "status": "ok",
            "file": str(filepath),
            "telegram_sent": sent
        }), 200

    except Exception as e:
        error_text = f"❌ Ошибка: {str(e)}"
        print(error_text)
        logging.error(error_text)
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route("/last_error", methods=["GET"])
def last_error():
    try:
        with open("errors.log", "r", encoding="utf-8") as f:
            return "<br>".join(f.readlines()[-10:]) or "Нет ошибок"
    except Exception as e:
        return f"Ошибка чтения лога: {str(e)}"

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)
